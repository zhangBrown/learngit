# -*- coding: utf-8 -*-
import sys, os
sys.path.append("/home/zhangweibin/api_auto")

from openpyxl import load_workbook
from public.Log_method import Logger
from public.login import get_pwd
import json


loger = Logger("oper_exl", file_dir="/home/zhangweibin/api_auto/data/test.log")
auth_token, token, deviceid = get_pwd()  # 获取登录参数


class OperExcel:
    """读写excel"""

    def __init__(self, filename="/home/zhangweibin/api_auto/data/testcase2.xlsx", sheetname="Sheet1"):

        self.filename = filename
        self.sheetname = sheetname

    def read(self):
        wb = load_workbook(self.filename)
        ws = wb[self.sheetname]
        if ws.title == "Sheet1":
            row_max = ws.max_row
            loger.debug("总共有%s条用例" % (row_max-1))
            if row_max > 1:

                list = []
                for i in range(2, row_max+1):
                        list.append({
                            "case_name": ws["B%d" % i].value,
                            "method": ws["C%d" % i].value,
                            "url": ws["D%d" % i].value,
                            "headers": ws["E%d" % i].value,
                            "data": ws["F%d" % i].value,
                            "expect_result": ws["G%d" % i].value,
                            # "other_result": ws["H%d" % i].value,
                        })
                for i in list:  # 追加登录参数
                    a = {"AUTH-TOKEN": auth_token}
                    a = json.dumps(a)
                    i["headers"] = a

                    j = i["data"]
                    j = eval(j)
                    j["token"] = token
                    j["device_id"] = deviceid
                    i["data"] = json.dumps(j)
                return list
            else:
                print("没数据啊，你醒醒吧")

    def write(self, row, col, values):

        wb = load_workbook(self.filename)
        ws = wb.active
        ws.cell(row=row, column=col).value = values
        wb.save(self.filename)

if __name__ == "__main__":
    wbs = OperExcel("/home/zhangweibin/api_auto/data/testcase2.xlsx").read()
    for i in wbs:
        print(i)








