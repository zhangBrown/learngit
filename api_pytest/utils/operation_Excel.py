from openpyxl import load_workbook




class OperExcel:
    """读写excel"""

    def __init__(self, filename=None, sheetname="Sheet1"):

        self.filename = filename
        self.sheetname = sheetname

    def read(self):
        wb = load_workbook(self.filename)
        ws = wb[self.sheetname]
        if ws.title == "Sheet1":
            row_max = ws.max_row
            if row_max > 1:

                list = []
                for i in range(2, row_max+1):
                    case_dit = {}
                    case_dit["case_name"] = ws["B%d" % i].value
                    case_dit["except"] = ws["G%d" % i].value
                    case_dit["api"] = {
                            "method": ws["C%d" % i].value,
                            "url": ws["D%d" % i].value,
                            "headers": ws["E%d" % i].value,
                            "params": ws["F%d" % i].value,
                    }
                    list.append(case_dit)
                return list
            else:
                print("u have no data")

    def write(self, row, col, values):

        wb = load_workbook(self.filename)
        ws = wb.active
        ws.cell(row=row, column=col).value = values
        wb.save(self.filename)


if __name__ == "__main__":
    wbs = OperExcel("/home/zhangweibin/api_pytest/data/xlsx/test_data.xlsx")
    data = wbs.read()
    print(data)







